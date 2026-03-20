"""Export routes for Axion API — CSV, Excel, PDF."""

import csv
import io
import logging
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.api.deps import get_session
from src.database.models import Holding as HoldingModel, Security, Event, Alert, Digest

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/export", tags=["export"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _portfolio_rows(rows):
    """Extract portfolio data rows from DB results."""
    headers = [
        "Ticker", "Name", "Sector", "Geography", "Currency",
        "Quantity", "Avg Cost", "Current Price", "Market Value",
        "Weight %", "P&L", "P&L %",
    ]
    data = []
    for h, s in rows:
        mv = h.market_value or ((h.current_price or 0) * (h.quantity or 0)) or 0
        cost = (h.quantity or 0) * (h.avg_cost_basis or 0)
        pnl = mv - cost if cost else 0
        pnl_pct = (pnl / cost * 100) if cost else 0
        data.append([
            h.ticker,
            s.name if s else "",
            s.sector if s else "",
            s.geography if s else "",
            h.currency,
            h.quantity,
            h.avg_cost_basis or "",
            h.current_price or "",
            round(mv, 2),
            round(h.weight_pct or 0, 2),
            round(pnl, 2),
            round(pnl_pct, 2),
        ])
    return headers, data


def _events_rows(rows):
    """Extract events data rows from DB results."""
    headers = ["ID", "Title", "Type", "Materiality", "Published", "URL", "Source ID"]
    data = []
    for e in rows:
        data.append([
            e.id, e.title, e.event_type or "", e.materiality or "",
            str(e.published_at or ""), e.url or "", e.source_id or "",
        ])
    return headers, data


def _csv_response(headers, data, filename):
    """Build a CSV StreamingResponse."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _excel_response(headers, data, filename, sheet_name="Sheet1"):
    """Build an Excel StreamingResponse using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Auto-width columns
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row in data:
            if col_idx <= len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _pdf_response(headers, data, filename, title="Axion Report"):
    """Build a PDF StreamingResponse using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise RuntimeError("reportlab not installed. Run: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()

    elements = []

    # Title
    title_style = styles["Title"]
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 8 * mm))

    # Table data
    table_data = [headers] + [[str(v) for v in row] for row in data]

    # Calculate column widths proportionally
    page_width = landscape(A4)[0] - 30 * mm
    n_cols = len(headers)
    col_width = page_width / n_cols

    t = Table(table_data, colWidths=[col_width] * n_cols, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B579A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@router.get("/portfolio")
async def export_portfolio(
    format: str = Query("csv", description="Export format: csv, xlsx, pdf"),
    session: AsyncSession = Depends(get_session),
):
    """Export portfolio holdings."""
    stmt = (
        select(HoldingModel, Security)
        .outerjoin(Security, HoldingModel.ticker == Security.ticker)
        .where(HoldingModel.status == "active")
        .order_by(HoldingModel.ticker)
    )
    rows = (await session.execute(stmt)).all()
    headers, data = _portfolio_rows(rows)

    fmt = format.lower()
    if fmt == "xlsx":
        return _excel_response(headers, data, "portfolio_export.xlsx", "Portfolio")
    elif fmt == "pdf":
        return _pdf_response(headers, data, "portfolio_export.pdf", "Portfolio Holdings")
    else:
        return _csv_response(headers, data, "portfolio_export.csv")


@router.get("/events")
async def export_events(
    format: str = Query("csv", description="Export format: csv, xlsx, pdf"),
    session: AsyncSession = Depends(get_session),
):
    """Export events."""
    stmt = select(Event).order_by(Event.published_at.desc()).limit(1000)
    rows = (await session.execute(stmt)).scalars().all()
    headers, data = _events_rows(rows)

    fmt = format.lower()
    if fmt == "xlsx":
        return _excel_response(headers, data, "events_export.xlsx", "Events")
    elif fmt == "pdf":
        return _pdf_response(headers, data, "events_export.pdf", "Event Log")
    else:
        return _csv_response(headers, data, "events_export.csv")


@router.get("/alerts")
async def export_alerts(
    format: str = Query("csv", description="Export format: csv, xlsx, pdf"),
    session: AsyncSession = Depends(get_session),
):
    """Export alerts."""
    stmt = select(Alert).order_by(Alert.created_at.desc()).limit(500)
    rows = (await session.execute(stmt)).scalars().all()

    headers = ["ID", "Severity", "Title", "Body", "Acknowledged", "Created"]
    data = []
    for a in rows:
        data.append([
            a.id, a.severity or "", a.title or "", a.body or "",
            "Yes" if a.acknowledged else "No", str(a.created_at or ""),
        ])

    fmt = format.lower()
    if fmt == "xlsx":
        return _excel_response(headers, data, "alerts_export.xlsx", "Alerts")
    elif fmt == "pdf":
        return _pdf_response(headers, data, "alerts_export.pdf", "Alert Report")
    else:
        return _csv_response(headers, data, "alerts_export.csv")


@router.get("/digest")
async def export_digest(
    format: str = Query("pdf", description="Export format: pdf, csv"),
    session: AsyncSession = Depends(get_session),
):
    """Export latest digest as PDF or CSV."""
    stmt = select(Digest).order_by(Digest.created_at.desc()).limit(1)
    result = (await session.execute(stmt)).scalar_one_or_none()

    if not result:
        return StreamingResponse(
            iter(["No digest available"]),
            media_type="text/plain",
            headers={"Content-Disposition": "attachment; filename=no_digest.txt"},
        )

    headers = ["Section", "Content"]
    data = []
    import json
    try:
        parsed = json.loads(result.content) if result.content else {}
    except (json.JSONDecodeError, TypeError):
        parsed = {}

    if isinstance(parsed, dict) and "sections" in parsed:
        for s in parsed["sections"]:
            title = s.get("title", "")
            content = s.get("content", "")
            if isinstance(content, (list, dict)):
                content = json.dumps(content, indent=2)
            data.append([title, str(content)])
    elif isinstance(parsed, dict):
        for k, v in parsed.items():
            if isinstance(v, (list, dict)):
                v = json.dumps(v, indent=2)
            data.append([k, str(v)])
    else:
        data.append(["Digest", result.content or "No content"])

    fmt = format.lower()
    if fmt == "pdf":
        return _pdf_response(headers, data, "digest_export.pdf", "Intelligence Digest")
    elif fmt == "xlsx":
        return _excel_response(headers, data, "digest_export.xlsx", "Digest")
    else:
        return _csv_response(headers, data, "digest_export.csv")
